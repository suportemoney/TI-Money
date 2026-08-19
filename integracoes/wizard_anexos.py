"""Extrai texto de anexos do wizard (OCR, PDF, CSV, XLSX) para o DeepSeek."""

from __future__ import annotations

import base64
import csv
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

MAX_ARQUIVOS = 4
MAX_BYTES = 5 * 1024 * 1024
MAX_TEXTO = 12000
MAX_LINHAS_PLANILHA = 80
MAX_COLUNAS = 24

_IMAGENS = {'.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp', '.tif', '.tiff'}
_PDF = {'.pdf'}
_CSV = {'.csv', '.txt'}
_XLSX = {'.xlsx'}
_XLS = {'.xls'}


def _ext(nome: str, mime: str) -> str:
    suf = Path(nome or '').suffix.lower()
    if suf:
        return suf
    mime = (mime or '').lower()
    if mime in ('image/png',):
        return '.png'
    if mime in ('image/jpeg', 'image/jpg'):
        return '.jpg'
    if mime in ('image/webp',):
        return '.webp'
    if mime in ('image/gif',):
        return '.gif'
    if mime in ('application/pdf',):
        return '.pdf'
    if mime in ('text/csv',):
        return '.csv'
    if 'spreadsheet' in mime or mime.endswith('sheet'):
        return '.xlsx'
    return ''


def _decodificar(item: dict) -> bytes:
    b64 = (item.get('data_base64') or '').strip()
    if not b64:
        raise ValueError('Arquivo sem conteúdo.')
    try:
        bruto = base64.b64decode(b64, validate=False)
    except Exception as exc:
        raise ValueError('Base64 inválido.') from exc
    if len(bruto) > MAX_BYTES:
        raise ValueError(f'Arquivo passa de {MAX_BYTES // (1024 * 1024)} MB.')
    return bruto


def _ler_csv(raw: bytes) -> str:
    for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            texto = raw.decode(enc)
            break
        except UnicodeDecodeError:
            texto = ''
    else:
        texto = raw.decode('utf-8', errors='replace')
    amostra = texto[:4096]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(texto), dialect)
    linhas = []
    for i, row in enumerate(reader):
        if i >= MAX_LINHAS_PLANILHA:
            linhas.append('[…linhas truncadas…]')
            break
        linhas.append('\t'.join(str(c)[:80] for c in row[:MAX_COLUNAS]))
    return '\n'.join(linhas)


def _ler_xlsx(raw: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('openpyxl não instalado no servidor.') from exc
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    partes = []
    try:
        for si, sheet in enumerate(wb.worksheets[:3]):
            partes.append(f'--- Aba {sheet.title} ---')
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= MAX_LINHAS_PLANILHA:
                    partes.append('[…linhas truncadas…]')
                    break
                valores = ['' if c is None else str(c)[:80] for c in row[:MAX_COLUNAS]]
                if any(v.strip() for v in valores):
                    partes.append('\t'.join(valores))
    finally:
        wb.close()
    return '\n'.join(partes)


def _ler_imagem(raw: bytes, nome: str) -> str:
    from integracoes.texto_local import extrair_texto_imagem_bytes, formatar_resultado_ocr

    ocr = extrair_texto_imagem_bytes(raw)
    bloco = formatar_resultado_ocr(ocr, origem='imagem')
    if ocr.strip():
        return bloco
    # Sem texto no OCR: tenta visão (ChatGPT/Gemini), se cadastrada
    try:
        from integracoes.llm import LlmError, chat_completion_vision, obter_integracao_visao

        if obter_integracao_visao():
            desc = chat_completion_vision(
                'Descreva este print de sistema em português: tabelas, nomes, '
                'ramais, logins, erros e o que o usuário provavelmente quer fazer.',
                raw,
            )
            if desc:
                return bloco + '\n\nDescrição (visão IA):\n' + desc.strip()
    except LlmError:
        pass
    except Exception:
        logger.exception('Visão IA falhou no anexo %s', nome)
    return bloco


def _ler_pdf(raw: bytes) -> str:
    from integracoes.texto_local import extrair_texto_pdf_bytes, formatar_resultado_ocr

    texto, metodo = extrair_texto_pdf_bytes(raw)
    return formatar_resultado_ocr(texto, origem=metodo or 'pdf_texto')


def _um_arquivo(item: dict) -> str:
    nome = (item.get('nome') or 'arquivo').strip() or 'arquivo'
    ext = _ext(nome, item.get('mime') or '')
    raw = _decodificar(item)
    if ext in _IMAGENS:
        corpo = _ler_imagem(raw, nome)
    elif ext in _PDF:
        corpo = _ler_pdf(raw)
    elif ext in _CSV:
        corpo = 'Planilha CSV:\n' + _ler_csv(raw)
    elif ext in _XLSX:
        corpo = 'Planilha XLSX:\n' + _ler_xlsx(raw)
    elif ext in _XLS:
        corpo = 'Arquivo .xls antigo não é lido. Salve como .xlsx ou .csv.'
    else:
        corpo = f'Tipo não suportado ({ext or "sem extensão"}). Use print, PDF, CSV ou XLSX.'
    return f'### Arquivo: {nome}\n{corpo}'


def extrair_anexos_wizard(anexos) -> str:
    """Junta o texto extraído dos anexos do POST. String vazia se não houver."""
    if not anexos or not isinstance(anexos, list):
        return ''
    blocos = []
    for item in anexos[:MAX_ARQUIVOS]:
        if not isinstance(item, dict):
            continue
        nome = item.get('nome') or 'arquivo'
        try:
            blocos.append(_um_arquivo(item))
        except Exception as exc:
            logger.exception('Falha ao ler anexo do wizard %s', nome)
            blocos.append(f'### Arquivo: {nome}\n[Não foi possível ler: {exc}]')
    if not blocos:
        return ''
    texto = '\n\n'.join(blocos)
    if len(texto) > MAX_TEXTO:
        texto = texto[: MAX_TEXTO - 40].rstrip() + '\n\n[…anexos truncados…]'
    return 'Arquivos anexados pelo gestor (OCR/planilha):\n' + texto
